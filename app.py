import asyncio
import atexit
import json
import os
import re
import secrets
import subprocess
import sys
from datetime import datetime
from functools import wraps

import httpx  # type: ignore
import requests  # type: ignore
from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    render_template_string,
    request,
    session,
    send_from_directory,
    url_for,
)  # type: ignore

from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.formatting.rule import CellIsRule  # type: ignore
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore

from reportlab.lib import colors  # type: ignore
from reportlab.lib.pagesizes import A4, landscape  # type: ignore
from reportlab.lib.units import mm  # type: ignore
from reportlab.lib.utils import simpleSplit  # type: ignore
from reportlab.pdfgen import canvas  # type: ignore

from config import (
    FLASK_SECRET_KEY, SESSION_LIFETIME_SECONDS,
    SSO_MODE, DEV_USER_EMAIL, DEV_DOCENTE_EMAIL,
    DEV_RSPP_EMAIL, DEV_DIRIGENTE_EMAIL, DEV_UFFICIO_TECNICO_EMAIL,
    SSO_JWT_SECRET, SSO_JWT_ISSUER, SSO_JWT_AUDIENCE, SSO_PORTAL_URL,
    MAX_SESSIONS_PER_USER, MAX_SESSIONS_GLOBAL,
    WHITELIST_FILE, API_TOKEN, DEBUG, SSO_CONFIG
)
from shared_modules.sso_middleware import SSOMiddleware, RateLimiter, RoleManager, render_sso_error


# ============================================================
# BOOTSTRAP APPLICAZIONE
# ============================================================

app = Flask(__name__)

app.secret_key = FLASK_SECRET_KEY
app.permanent_session_lifetime = SESSION_LIFETIME_SECONDS
app.debug = DEBUG

# Inizializza role manager (assegna ruoli in base alla forma dell'email e whitelist.json)
role_manager = RoleManager(WHITELIST_FILE)

# Inizializza rate limiter
rate_limiter = RateLimiter(
    max_sessions_per_user=MAX_SESSIONS_PER_USER,
    max_sessions_global=MAX_SESSIONS_GLOBAL,
    session_ttl_seconds=SESSION_LIFETIME_SECONDS
)

sso_middleware = SSOMiddleware(
    jwt_secret=SSO_JWT_SECRET,
    jwt_algorithm="HS256",
    jwt_issuer=SSO_JWT_ISSUER,
    jwt_audience=SSO_JWT_AUDIENCE,
    session_timeout=SESSION_LIFETIME_SECONDS,
    portal_url=SSO_PORTAL_URL,
    rate_limiter=rate_limiter
)


# ============================================================
# DATI STATICI
# ============================================================

with open("floors.json", "r", encoding="utf-8") as f:
    aule_dict = json.load(f)

aula = []
for _, rooms in aule_dict.items():
    aula.extend(rooms)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REGISTRI_EXPORT_DIR = os.path.join(BASE_DIR, "exports", "registri_compilati")
REGISTRI_EXPORT_PDF_DIR = os.path.join(REGISTRI_EXPORT_DIR, "pdf")


def _format_date_value(value):
    if not value:
        return ""

    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return value


def _format_time_value(value):
    if not value:
        return ""

    try:
        return datetime.strptime(value, "%H:%M:%S").strftime("%H:%M")
    except ValueError:
        return value


def _build_registri_export_filename(classe_key):
    safe_classe_key = re.sub(r"[^A-Za-z0-9_-]+", "_", str(classe_key).strip().upper())
    return f"registro_{safe_classe_key}.xlsx"


def _load_registri_per_classe():
    import sqlite3

    registri_per_classe = {}
    compilazioni_per_classe = {}

    with sqlite3.connect(os.path.join(BASE_DIR, "runout.db")) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT 
                c.ClasseID,
                c.Anno,
                c.Sezione,
                r.StudenteID,
                r.Nome,
                r.Cognome,
                r.Stato,
                s.NomeStato,
                s.Descrizione
            FROM Registri r
            JOIN Classi c ON r.Classe = c.ClasseID
            LEFT JOIN Stati s ON r.Stato = s.StatoID
            ORDER BY c.Anno ASC, c.Sezione ASC, r.Cognome ASC, r.Nome ASC
            """
        )
        rows = cursor.fetchall()

        cursor.execute(
            """
            SELECT sl1.Classe, sl1.Data, sl1.Ora, sl1.Timestamp
            FROM SyncLog sl1
            WHERE sl1.Timestamp = (
                SELECT MAX(sl2.Timestamp)
                FROM SyncLog sl2
                WHERE sl2.Classe = sl1.Classe
            )
            """
        )
        sync_rows = cursor.fetchall()

    for row in sync_rows:
        classe_key = str(row["Classe"]).strip().upper()
        compilazioni_per_classe[classe_key] = {
            "data": _format_date_value(row["Data"] or ""),
            "ora": _format_time_value(row["Ora"] or ""),
            "timestamp": row["Timestamp"] or "",
        }

    for row in rows:
        classe_key = f"{row['Anno']}{row['Sezione']}"

        if classe_key not in registri_per_classe:
            registri_per_classe[classe_key] = {
                "anno": row["Anno"],
                "sezione": row["Sezione"],
                "studenti": [],
                "data_compilazione": compilazioni_per_classe.get(classe_key, {}).get("data", ""),
                "ora_compilazione": compilazioni_per_classe.get(classe_key, {}).get("ora", ""),
            }

        studente = {
            "nome": row["Nome"] or "",
            "cognome": row["Cognome"] or "",
            "stato": row["NomeStato"] or "Non definito",
            "stato_id": row["Stato"],
            "descrizione": row["Descrizione"] or "",
        }
        registri_per_classe[classe_key]["studenti"].append(studente)

        if not registri_per_classe[classe_key].get("data_compilazione"):
            registri_per_classe[classe_key]["data_compilazione"] = compilazioni_per_classe.get(classe_key, {}).get("data", "")
        if not registri_per_classe[classe_key].get("ora_compilazione"):
            registri_per_classe[classe_key]["ora_compilazione"] = compilazioni_per_classe.get(classe_key, {}).get("ora", "")

    return dict(sorted(registri_per_classe.items(), key=lambda x: (x[1]["anno"], x[1]["sezione"])))


def _load_compilazioni_per_classe():
    presenze_path = os.path.join(BASE_DIR, "presenze.json")
    compilazioni_per_classe = {}

    if not os.path.exists(presenze_path):
        return compilazioni_per_classe

    try:
        with open(presenze_path, "r", encoding="utf-8") as file_handle:
            records = json.load(file_handle)
    except (OSError, json.JSONDecodeError):
        return compilazioni_per_classe

    if not isinstance(records, list):
        return compilazioni_per_classe

    for record in records:
        classe_key = str(record.get("classe", "")).strip().upper()
        if not classe_key:
            continue

        timestamp_value = str(record.get("timestamp", "")).strip()
        current = compilazioni_per_classe.get(classe_key)
        if current and current.get("timestamp", "") >= timestamp_value:
            continue

        compilazioni_per_classe[classe_key] = {
            "data": _format_date_value(record.get("data", "")),
            "ora": _format_time_value(record.get("ora", "")),
            "docente": (record.get("docente_email") or record.get("docente") or "Non disponibile"),
            "timestamp": timestamp_value,
        }

    return compilazioni_per_classe


def _stato_lettera(stato):
    stato_norm = str(stato or "").strip().upper()
    if stato_norm in {"P", "PRESENTE"}:
        return "P"
    if stato_norm in {"A", "ASSENTE"}:
        return "A"
    if stato_norm in {"D", "DISPERSO"}:
        return "D"
    return ""


def _generate_registri_excel_exports(registri_per_classe):
    os.makedirs(REGISTRI_EXPORT_DIR, exist_ok=True)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="FFFFFF")
    info_label_fill = PatternFill("solid", fgColor="E8EEF7")
    info_value_fill = PatternFill("solid", fgColor="F8FAFC")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    export_info = {}
    for classe_key, classe_data in registri_per_classe.items():
        filename = _build_registri_export_filename(classe_key)
        file_path = os.path.join(REGISTRI_EXPORT_DIR, filename)

        classe_label = f"{classe_data['anno']}{classe_data['sezione']}"
        studenti = classe_data.get("studenti", [])
        compilazione = classe_data.get("compilazione", {})
        docente_email = compilazione.get("docente", "Non disponibile")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Registro"

        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A4"
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1
        worksheet.page_setup.orientation = "portrait"
        worksheet.page_margins.left = 0.28
        worksheet.page_margins.right = 0.28
        worksheet.page_margins.top = 0.4
        worksheet.page_margins.bottom = 0.4
        worksheet.print_options.horizontalCentered = True

        worksheet.merge_cells("A1:B1")
        worksheet.merge_cells("A2:B2")
        worksheet["A1"] = "ITIS P. Paleocapa"
        worksheet["A2"] = f"APPELLO CLASSE {classe_label}"

        title_cell = worksheet["A1"]
        title_cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = title_fill

        subtitle_cell = worksheet["A2"]
        subtitle_cell.font = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        subtitle_cell.fill = subtitle_fill

        header_row = 4
        worksheet["A4"] = "STUDENTE"
        worksheet["B4"] = "STATO"
        for cell in worksheet[header_row]:
            cell.font = Font(name="Calibri", size=10, bold=True, color="1F1F1F")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        data_start_row = 5
        for index, studente in enumerate(studenti, start=data_start_row):
            worksheet[f"A{index}"] = f"{studente['cognome']} {studente['nome']}".strip()
            worksheet[f"B{index}"] = _stato_lettera(studente["stato"])

            worksheet[f"A{index}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            worksheet[f"B{index}"].alignment = Alignment(horizontal="center", vertical="center")

            if worksheet[f"B{index}"].value == "P":
                worksheet[f"B{index}"].fill = green_fill
            elif worksheet[f"B{index}"].value == "A":
                worksheet[f"B{index}"].fill = yellow_fill
            elif worksheet[f"B{index}"].value == "D":
                worksheet[f"B{index}"].fill = red_fill

            worksheet[f"A{index}"].border = thin_border
            worksheet[f"B{index}"].border = thin_border

        if studenti:
            worksheet.conditional_formatting.add(
                f"B{data_start_row}:B{worksheet.max_row}",
                CellIsRule(operator="equal", formula=['"P"'], fill=green_fill),
            )
            worksheet.conditional_formatting.add(
                f"B{data_start_row}:B{worksheet.max_row}",
                CellIsRule(operator="equal", formula=['"A"'], fill=yellow_fill),
            )
            worksheet.conditional_formatting.add(
                f"B{data_start_row}:B{worksheet.max_row}",
                CellIsRule(operator="equal", formula=['"D"'], fill=red_fill),
            )
        else:
            worksheet[f"A{data_start_row}"] = "Nessuno studente registrato"
            worksheet[f"B{data_start_row}"] = ""
            worksheet[f"A{data_start_row}"].border = thin_border
            worksheet[f"B{data_start_row}"].border = thin_border

        info_row = worksheet.max_row + 2
        info_items = [
            ("DATA", classe_data.get("data_compilazione", "")),
            ("ORA", classe_data.get("ora_compilazione", "")),
            ("DOCENTE", docente_email),
        ]

        for offset, (label, value) in enumerate(info_items):
            row_number = info_row + offset
            worksheet[f"A{row_number}"] = label
            worksheet[f"B{row_number}"] = value or "Non disponibile"
            worksheet[f"A{row_number}"].font = Font(name="Calibri", size=9, bold=True)
            worksheet[f"A{row_number}"].fill = info_label_fill
            worksheet[f"A{row_number}"].border = thin_border
            worksheet[f"A{row_number}"].alignment = Alignment(horizontal="left", vertical="center")
            worksheet[f"B{row_number}"].font = Font(name="Calibri", size=9)
            worksheet[f"B{row_number}"].fill = info_value_fill
            worksheet[f"B{row_number}"].border = thin_border
            worksheet[f"B{row_number}"].alignment = Alignment(horizontal="left", vertical="center")

        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[2].height = 18
        worksheet.row_dimensions[4].height = 18

        worksheet.column_dimensions["A"].width = 48
        worksheet.column_dimensions["B"].width = 12

        workbook.save(file_path)
        export_info[classe_key] = {
            "filename": filename,
        }

    return export_info


def _pdf_filename_for_classe(classe_key):
    safe_classe_key = re.sub(r"[^A-Za-z0-9_-]+", "_", str(classe_key).strip().upper())
    return f"registro_{safe_classe_key}.pdf"


def _generate_registri_pdf_exports(registri_per_classe):
    os.makedirs(REGISTRI_EXPORT_PDF_DIR, exist_ok=True)

    page_width, page_height = A4
    margin_left = 10 * mm
    margin_right = 10 * mm
    margin_top = 10 * mm
    margin_bottom = 10 * mm
    footer_height = 24 * mm
    usable_width = page_width - margin_left - margin_right
    student_col_width = usable_width * 0.83
    state_col_width = usable_width * 0.17
    title_top_y = page_height - margin_top
    subtitle_top_y = title_top_y - 7 * mm
    table_top_y = subtitle_top_y - 12 * mm
    footer_top_limit = margin_bottom + footer_height

    export_info = {}
    for classe_key, classe_data in registri_per_classe.items():
        filename = _pdf_filename_for_classe(classe_key)
        file_path = os.path.join(REGISTRI_EXPORT_PDF_DIR, filename)

        classe_label = f"{classe_data['anno']}{classe_data['sezione']}"
        studenti = classe_data.get("studenti", [])
        compilazione = classe_data.get("compilazione", {})
        data_compilazione = compilazione.get("data") or classe_data.get("data_compilazione", "Non disponibile")
        ora_compilazione = compilazione.get("ora") or classe_data.get("ora_compilazione", "Non disponibile")
        docente_email = compilazione.get("docente", "Non disponibile")
        pdf_canvas = canvas.Canvas(file_path, pagesize=A4)
        pdf_canvas.setTitle(f"Registro classe {classe_label}")
        pdf_canvas.setAuthor("Runout")

        def draw_header():
            pdf_canvas.setFillColor(colors.HexColor("#1F1F1F"))
            pdf_canvas.setFont("Helvetica-Bold", 18)
            pdf_canvas.drawCentredString(page_width / 2, title_top_y, "ITIS P. Paleocapa")
            pdf_canvas.setFont("Helvetica-Bold", 11)
            pdf_canvas.drawCentredString(page_width / 2, subtitle_top_y, f"APPELLO CLASSE {classe_label}")

            header_y = table_top_y
            pdf_canvas.setFillColor(colors.white)
            pdf_canvas.rect(margin_left, header_y - 7 * mm, student_col_width, 7 * mm, fill=1, stroke=1)
            pdf_canvas.rect(margin_left + student_col_width, header_y - 7 * mm, state_col_width, 7 * mm, fill=1, stroke=1)
            pdf_canvas.setFillColor(colors.HexColor("#1F1F1F"))
            pdf_canvas.setFont("Helvetica-Bold", 10)
            pdf_canvas.drawCentredString(margin_left + student_col_width / 2, header_y - 4.6 * mm, "STUDENTE")
            pdf_canvas.drawCentredString(margin_left + student_col_width + state_col_width / 2, header_y - 4.6 * mm, "STATO")

        draw_header()

        current_y = table_top_y - 7 * mm
        if not studenti:
            studenti = [{"cognome": "Nessuno studente registrato", "nome": "", "stato": ""}]

        remaining_height = max(current_y - footer_top_limit, 40 * mm)
        row_height = remaining_height / max(len(studenti), 1)
        student_font_size = 9.5 if row_height >= 8.0 * mm else 8.5 if row_height >= 6.5 * mm else 7.5
        student_leading = student_font_size + 1.0

        for studente in studenti:
            stato_lettera = _stato_lettera(studente["stato"])
            student_text = f"{studente['cognome']} {studente['nome']}".strip()
            wrapped_lines = simpleSplit(student_text, "Helvetica", student_font_size, student_col_width - 4 * mm)

            if stato_lettera == "P" or stato_lettera == "A" or stato_lettera == "D":
                fill_map = {
                    "P": colors.HexColor("#C6EFCE"),
                    "A": colors.HexColor("#FFEB9C"),
                    "D": colors.HexColor("#FFC7CE"),
                }
                state_fill = fill_map[stato_lettera]
            else:
                state_fill = colors.white

            pdf_canvas.setStrokeColor(colors.black)
            pdf_canvas.setFillColor(colors.white)
            pdf_canvas.rect(margin_left, current_y - row_height, student_col_width, row_height, fill=1, stroke=1)
            pdf_canvas.setFillColor(state_fill)
            pdf_canvas.rect(margin_left + student_col_width, current_y - row_height, state_col_width, row_height, fill=1, stroke=1)

            pdf_canvas.setFillColor(colors.HexColor("#1F1F1F"))
            pdf_canvas.setFont("Helvetica", student_font_size)
            text_y = current_y - (row_height / 2) + ((len(wrapped_lines) - 1) * student_leading / 2)
            for line in wrapped_lines:
                pdf_canvas.drawString(margin_left + 2 * mm, text_y, line)
                text_y -= student_leading

            pdf_canvas.setFont("Helvetica-Bold", 10)
            pdf_canvas.drawCentredString(margin_left + student_col_width + state_col_width / 2, current_y - row_height / 2 - 1.2 * mm, stato_lettera)

            current_y -= row_height

        footer_y = margin_bottom + 16 * mm
        pdf_canvas.setFont("Helvetica-Bold", 9)
        footer_items = [
            ("DATA", data_compilazione or "Non disponibile"),
            ("ORA", ora_compilazione or "Non disponibile"),
            ("DOCENTE", docente_email or "Non disponibile"),
        ]
        for index, (label, value) in enumerate(footer_items):
            y_position = footer_y - (index * 5 * mm)
            pdf_canvas.setFillColor(colors.HexColor("#1F1F1F"))
            pdf_canvas.drawString(margin_left, y_position, f"{label}")
            pdf_canvas.setFont("Helvetica", 9)
            pdf_canvas.drawString(margin_left + 15 * mm, y_position, value)
            pdf_canvas.setFont("Helvetica-Bold", 9)

        pdf_canvas.save()
        export_info[classe_key] = {"filename": filename}

    return export_info

# ============================================================
# CACHE EMERGENZE
# ============================================================

_emergenze_cache = {
    "risultati_filtrati": [],
    "giorno": None,
    "ora": None,
    "total_aule": 0,
    "num_with_class": 0,
    "last_update": None,
    "loading": False,
}
CACHE_TTL_SECONDS = 300  # Aggiorna la cache ogni 5 minuti


async def _fetch_emergenze_data():
    """Recupera i dati delle emergenze dalle API e aggiorna la cache in memoria."""
    global _emergenze_cache

    if _emergenze_cache["loading"]:
        return  # Evita fetch paralleli
    _emergenze_cache["loading"] = True

    try:
        now = datetime.now()
        giorno = now.weekday() + 1  # Lunedì=1 … Venerdì=5 (per API)
        
        # Nomi dei giorni della settimana
        giorni_nomi = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica']
        giorno_nome = giorni_nomi[now.weekday()]  # Nome del giorno della settimana
        
        ora_reale = now.hour
        ora = (ora_reale - 7) if 8 <= ora_reale <= 14 else 1

        async def fetch_classe(client, a):
            url = f"https://sipal.itispaleocapa.it/api/proxySipal/v1/studenti/classe/{giorno}/{ora}/{a}"
            headers = {
                "accept": "application/json",
                "Authorization": f"Bearer {API_TOKEN}",
                "User-Agent": "Mozilla/5.0",
            }
            try:
                response = await client.get(url, headers=headers)
                response.raise_for_status()
                return {"aula": a, "risultato": response.json(), "errore": None}
            except httpx.HTTPStatusError as e:
                return {"aula": a, "risultato": None, "errore": f"Errore HTTP {e.response.status_code}"}
            except Exception as e:
                return {"aula": a, "risultato": None, "errore": f"Errore generico: {str(e)}"}

        async with httpx.AsyncClient() as client:
            tasks = [fetch_classe(client, a) for a in aula]
            risultati = await asyncio.gather(*tasks)

        risultati_filtrati = []
        num_with_class = 0
        for r in risultati:
            classe = r["aula"]
            has_class = False
            if r["risultato"] and isinstance(r["risultato"], dict):
                if r["risultato"].get("classe"):
                    classe = r["risultato"]["classe"]
                    has_class = True
                elif r["risultato"].get("studenti") and len(r["risultato"]["studenti"]) > 0:
                    primo = r["risultato"]["studenti"][0]
                    if isinstance(primo, dict) and primo.get("classe"):
                        classe = primo["classe"]
                        has_class = True
            r["classe"] = classe
            if has_class:
                num_with_class += 1
            if classe != r["aula"]:
                risultati_filtrati.append(r)

        risultati_filtrati.sort(key=lambda x: x["classe"])

        _emergenze_cache.update({
            "risultati_filtrati": risultati_filtrati,
            "giorno": giorno,
            "giorno_nome": giorno_nome,
            "ora": ora,
            "total_aule": len(aula),
            "num_with_class": num_with_class,
            "last_update": datetime.now(),
            "loading": False,
        })
        app.logger.info(f"Cache emergenze aggiornata: {num_with_class} classi trovate")

    except Exception as e:
        _emergenze_cache["loading"] = False
        app.logger.error(f"Errore aggiornamento cache emergenze: {e}")


def _cache_is_stale() -> bool:
    last = _emergenze_cache["last_update"]
    if last is None:
        return True
    return (datetime.now() - last).total_seconds() > CACHE_TTL_SECONDS


def _refresh_cache_in_background():
    """Lancia il refresh della cache in un thread separato senza bloccare la risposta."""
    import threading

    def run():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(_fetch_emergenze_data())
        finally:
            loop.close()

    threading.Thread(target=run, daemon=True).start()


# ============================================================
# GESTIONE EMERGENZE - STATO ATTIVO/INATTIVO
# ============================================================

EMERGENZA_STATUS_FILE = 'emergenza_status.json'

def _get_emergenza_status():
    """Legge lo stato dell'emergenza dal file."""
    try:
        if os.path.exists(EMERGENZA_STATUS_FILE):
            with open(EMERGENZA_STATUS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        app.logger.error(f"Errore lettura stato emergenza: {e}")
    
    # Default se file non esiste o errore
    return {"active": False, "started_at": None, "ended_at": None}


def _set_emergenza_status(active: bool):
    """Imposta lo stato dell'emergenza nel file."""
    try:
        status = {
            "active": active,
            "started_at": datetime.now().isoformat() if active else None,
            "ended_at": datetime.now().isoformat() if not active else None,
        }
        with open(EMERGENZA_STATUS_FILE, 'w', encoding='utf-8') as f:
            json.dump(status, f, ensure_ascii=False, indent=2)
        app.logger.info(f"Stato emergenza: {'ATTIVA' if active else 'TERMINATA'}")
        return status
    except Exception as e:
        app.logger.error(f"Errore scrittura stato emergenza: {e}")
        return _get_emergenza_status()


# ============================================================
# UTILITY & DECORATORS
# ============================================================

def get_username(email: str) -> str:
    return email.split('@')[0]


def get_dev_roles():
    """
    Ritorna i 5 ruoli disponibili in modalità DEV con le loro email.
    """
    return {
        'studente': DEV_USER_EMAIL,
        'docente': DEV_DOCENTE_EMAIL,
        'rspp': DEV_RSPP_EMAIL,
        'dirigente': DEV_DIRIGENTE_EMAIL,
        'ufficio_tecnico': DEV_UFFICIO_TECNICO_EMAIL,
    }


def role_required(allowed_roles):
    """
    Decorator per proteggere le rotte in base al ruolo dell'utente.
    
    Uso:
        @role_required('docente')
        def my_route():
            ...
    """
    if isinstance(allowed_roles, str):
        allowed_roles = [allowed_roles]
    
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = session.get("user", None)
            if not user:
                app.logger.warning("Accesso senza sessione")
                return redirect(url_for("home"))
            
            user_role = user.get("role", "guest")
            if user_role not in allowed_roles:
                app.logger.warning(f"Accesso vietato per ruolo '{user_role}' a {request.path}")
                return render_template_string(
                    """
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="UTF-8">
                        <title>Accesso Vietato</title>
                        <style>
                            body { font-family: Arial; background: #f5f5f5; padding: 20px; }
                            .error-container { max-width: 600px; margin: 100px auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); text-align: center; }
                            .error-icon { font-size: 64px; margin-bottom: 20px; }
                            h1 { color: #d32f2f; margin: 0 0 12px; }
                            p { color: #666; margin: 12px 0; }
                            .user-role { background: #f0f0f0; padding: 12px; border-radius: 6px; margin: 20px 0; font-family: monospace; }
                            a { display: inline-block; margin-top: 20px; padding: 10px 24px; background: #667eea; color: white; text-decoration: none; border-radius: 6px; }
                            a:hover { background: #5568d3; }
                        </style>
                    </head>
                    <body>
                        <div class="error-container">
                            <div class="error-icon">🚫</div>
                            <h1>Accesso Vietato</h1>
                            <p>Non hai il permesso di accedere a questa pagina.</p>
                            <p>Ruoli consentiti: <strong>{{ allowed }}</strong></p>
                            <div class="user-role">Il tuo ruolo: <strong>{{ your_role }}</strong></div>
                            <a href="{{ home_url }}">← Torna alla Home</a>
                        </div>
                    </body>
                    </html>
                    """,
                    allowed=", ".join(allowed_roles),
                    your_role=user_role,
                    home_url=url_for("home")
                ), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ============================================================
# ROUTE PUBBLICHE
# ============================================================

@app.route("/")
def home():
    """Home page - mostra scelta ruoli in dev mode, oppure home normale."""
    user = session.get('user', None)
    
    # In dev mode, mostra direttamente la scelta di ruoli se non autenticato
    if SSO_MODE == 'dev' and not user:
        dev_roles = get_dev_roles()
        return render_template("dev_login_choice.html", dev_roles=dev_roles)
    
    return render_template("home.html", user=user, sso_mode=SSO_MODE)


# ============================================================
# ROUTE DEV
# ============================================================

@app.route("/dev/login-choice")
def dev_login_choice():
    """
    (Solo in DEV mode) Pagina per scegliere quale ruolo usare per il login di test.
    """
    if SSO_MODE != 'dev':
        return redirect(url_for('home'))
    
    dev_roles = get_dev_roles()
    return render_template("dev_login_choice.html", dev_roles=dev_roles)


# ============================================================
# ROUTE SSO E SESSIONE
# ============================================================

@app.route("/sso/login")
def sso_login():
    """
    Endpoint SSO. Il portale checkin chiama questa URL passando il JWT.
    In DEV mode, accetta il parametro role per simulare i diversi ruoli.
    Questo è l'unico punto di ingresso autenticato nell'applicazione.
    """
    token = request.args.get("token")

    # --- Modalità DEV: simula il login senza portale reale ---
    if SSO_MODE == 'dev' and not token:
        dev_role = request.args.get('role', 'studente')
        dev_roles = get_dev_roles()
        
        if dev_role not in dev_roles:
            return render_sso_error(
                f"Ruolo DEV non valido: {dev_role}. Ruoli disponibili: {', '.join(dev_roles.keys())}",
                SSO_CONFIG['portal_url'],
                status_code=400
            )
        
        dev_email = dev_roles[dev_role]
        app.logger.info(f"DEV MODE: login simulato per ruolo '{dev_role}' con email {dev_email}")
        user_data = {
            "email": dev_email,
            "name": get_username(dev_email).replace(".", " ").title(),
            "googleId": "dev-user-id",
            "picture": "",
        }
        return _complete_login(user_data)

    if not token:
        return render_sso_error(
            "Token SSO mancante. Accedi tramite il portale.",
            SSO_CONFIG['portal_url']
        )

    try:
        user_data = sso_middleware.validate_jwt(token)
        return _complete_login(user_data)
    except Exception as e:
        app.logger.error(f"Errore validazione SSO: {e}")
        return render_sso_error(
            f"Token SSO non valido o scaduto. Effettua nuovamente il login.",
            SSO_CONFIG['portal_url']
        )


def _complete_login(user_data: dict):
    """
    Logica comune post-validazione JWT:
    1. Verifica whitelist
    2. Verifica rate limit
    3. Crea sessione e redirect alla dashboard
    """
    email = user_data.get("email", "")

    role, is_authorized = role_manager.get_role(email)
    if not is_authorized:
        app.logger.warning(f"Accesso negato - Utente non autorizzato: {email} (role: {role})")
        return render_sso_error(
            f"Il tuo account ({email}) non è autorizzato ad accedere a questa applicazione. "
            "Contatta l'amministratore se ritieni sia un errore.",
            SSO_CONFIG['portal_url'],
            status_code=403,
            title="Account Non Autorizzato",
            icon="🚫"
        )
    
    app.logger.info(f"Utente autorizzato: {email} con ruolo '{role}'")

    session_id = secrets.token_hex(32)
    allowed, reason = rate_limiter.register_session(session_id, email)
    if not allowed:
        app.logger.warning(f"Rate limit raggiunto per: {email}")
        return render_sso_error(
            reason,
            SSO_CONFIG['portal_url'],
            status_code=429,
            title="Troppe Sessioni Attive",
            icon="⏱️"
        )

    sso_middleware.create_session(user_data, session, session_id=session_id, role=role)

    if _cache_is_stale():
        _refresh_cache_in_background()

    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    """Logout - termina la sessione e reindirizza al portale SSO o home."""
    if session.get("session_id"):
        rate_limiter.remove_session(session.get("session_id"))
    session.clear()
    app.logger.info("Logout effettuato")
    return redirect(url_for("home"))


# ============================================================
# ROUTE DOCENTE
# ============================================================

@app.route("/emergenze") # ROTTA EMERGENZE
@role_required(['docente', 'rspp', 'dirigente', 'ufficio_tecnico'])  # Docenti e staff
def emergenze():
    if _emergenze_cache["last_update"] is None:
        import time

        _refresh_cache_in_background()
        for _ in range(20):
            if _emergenze_cache["last_update"] is not None:
                break
            time.sleep(0.5)
    elif _cache_is_stale():
        _refresh_cache_in_background()

    c = _emergenze_cache
    emergenza_status = _get_emergenza_status()
    user = session.get('user', {})
    user_role = user.get('role', '')
    
    # Se emergenza non attiva, nessuno vede l'elenco delle classi
    if not emergenza_status["active"]:
        return render_template(
            "emergenze.html",
            risultati=[],
            giorno=c["giorno"],
            giorno_nome=c["giorno_nome"],
            ora=c["ora"],
            total_aule=0,
            num_with_class=0,
            emergenza_attiva=emergenza_status["active"],
            is_ufficio_tecnico=(user_role == 'ufficio_tecnico'),
        )
    
    return render_template(
        "emergenze.html",
        risultati=c["risultati_filtrati"],
        giorno=c["giorno"],
        giorno_nome=c["giorno_nome"],
        ora=c["ora"],
        total_aule=c["total_aule"],
        num_with_class=c["num_with_class"],
        emergenza_attiva=emergenza_status["active"],
        is_ufficio_tecnico=(user_role == 'ufficio_tecnico'),
    )


@app.route("/api/emergenze/refresh", methods=["POST"])
@role_required(['docente', 'rspp', 'dirigente', 'ufficio_tecnico'])  # Docenti e staff
def refresh_emergenze():
    """Forza il ricalcolo della cache emergenze (es. da un pulsante nella UI)."""
    _refresh_cache_in_background()
    return jsonify({"success": True, "message": "Aggiornamento cache avviato"}), 202


@app.route("/elencoStudenti/<classe>/<aula>") # ROTTA ELENCO STUDENTI
@role_required(['docente', 'rspp', 'dirigente', 'ufficio_tecnico'])  # Docenti e staff
def elencoStudenti(classe, aula):
    # Verifica che l'emergenza sia attiva
    emergenza_status = _get_emergenza_status()
    if not emergenza_status["active"]:
        return redirect(url_for('emergenze'))
    
    url = f"https://sipal.itispaleocapa.it/api/proxySipal/v1/studenti/classe/elenco/{classe}"
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {API_TOKEN}",
        "User-Agent": "Mozilla/5.0"
    }
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        dati_classe = response.json()
            
        studenti = []
        if isinstance(dati_classe, dict):
            if classe in dati_classe and isinstance(dati_classe.get(classe), list):
                studenti = dati_classe.get(classe) or []
            
        return render_template(
            "elenco_studenti.html",
            classe=classe,
            aula=aula,
            studenti=studenti,
            giorno=None,
            ora=None,
            errore=None,
            emergenza_attiva=emergenza_status["active"]
        )
    except requests.HTTPError as e:
        return render_template(
            "elenco_studenti.html",
            classe=classe,
            aula=aula,
            studenti=[],
            giorno=None,
            ora=None,
            errore=f"Errore HTTP {e.response.status_code}",
            emergenza_attiva=emergenza_status["active"]
        )
    except Exception as e:
        return render_template(
            "elenco_studenti.html",
            classe=classe,
            aula=aula,
            studenti=[],
            giorno=None,
            ora=None,
            errore=f"Errore generico: {str(e)}",
            emergenza_attiva=emergenza_status["active"]
        )

@app.route("/piantina") # ROTTA PIANTINA
@sso_middleware.sso_login_required  # Accessibile a tutti i ruoli loggati
def piantina():
    return render_template("piantina.html", pdf_path="/static/piantina.pdf")

@app.route("/registri-compilati") # ROTTA REGISTRI COMPILATI
@role_required(['rspp', 'dirigente', 'ufficio_tecnico'])  # Solo staff
def registri_compilati():
    try:
        registri_per_classe = _load_registri_per_classe()
        compilazioni_per_classe = _load_compilazioni_per_classe()

        for classe_key, classe_data in registri_per_classe.items():
            classe_data["compilazione"] = compilazioni_per_classe.get(classe_key, {})
        export_info = _generate_registri_excel_exports(registri_per_classe)
        export_info_pdf = _generate_registri_pdf_exports(registri_per_classe)

        for classe_key, classe_data in registri_per_classe.items():
            classe_data["download_filename"] = export_info.get(classe_key, {}).get("filename", _build_registri_export_filename(classe_key))
            classe_data["download_url"] = export_info.get(classe_key, {}).get(
                "download_url",
                url_for("download_registro_compilato", classe_key=classe_key),
            )
            classe_data["download_pdf_filename"] = export_info_pdf.get(classe_key, {}).get("filename", _pdf_filename_for_classe(classe_key))
            classe_data["download_pdf_url"] = url_for("download_registro_compilato_pdf", classe_key=classe_key)
    except Exception as e:
        print(f"Errore nel recupero dei registri: {e}")
        registri_per_classe = {}
    
    return render_template("registri_compilati.html", registri_per_classe=registri_per_classe)


@app.route("/registri-compilati/download/<classe_key>")
@role_required("docente")
def download_registro_compilato(classe_key):
    try:
        registri_per_classe = _load_registri_per_classe()
        compilazioni_per_classe = _load_compilazioni_per_classe()

        for classe_key_corrente, classe_data in registri_per_classe.items():
            classe_data["compilazione"] = compilazioni_per_classe.get(classe_key_corrente, {})

        if classe_key not in registri_per_classe:
            return "Registro non trovato", 404

        _generate_registri_excel_exports(registri_per_classe)
        filename = _build_registri_export_filename(classe_key)
        file_path = os.path.join(REGISTRI_EXPORT_DIR, filename)

        if not os.path.exists(file_path):
            return "Registro non trovato", 404

        return send_from_directory(REGISTRI_EXPORT_DIR, filename, as_attachment=True)
    except Exception as e:
        print(f"Errore nel download del registro {classe_key}: {e}")
        return "Errore nella generazione del registro", 500


@app.route("/registri-compilati/download/<classe_key>/pdf")
@role_required("docente")
def download_registro_compilato_pdf(classe_key):
    try:
        registri_per_classe = _load_registri_per_classe()
        compilazioni_per_classe = _load_compilazioni_per_classe()

        for classe_key_corrente, classe_data in registri_per_classe.items():
            classe_data["compilazione"] = compilazioni_per_classe.get(classe_key_corrente, {})

        if classe_key not in registri_per_classe:
            return "Registro non trovato", 404

        _generate_registri_pdf_exports(registri_per_classe)
        filename = _pdf_filename_for_classe(classe_key)
        file_path = os.path.join(REGISTRI_EXPORT_PDF_DIR, filename)

        if not os.path.exists(file_path):
            return "Registro non trovato", 404

        return send_from_directory(REGISTRI_EXPORT_PDF_DIR, filename, as_attachment=True)
    except Exception as e:
        print(f"Errore nel download PDF del registro {classe_key}: {e}")
        return "Errore nella generazione del registro PDF", 500


# ============================================================
# ROUTE API
# ============================================================

@app.route("/api/emergenze", methods=["POST"])
@role_required(['docente', 'rspp', 'dirigente', 'ufficio_tecnico'])  # Docenti e staff
def salva_presenze():
    """
    Salva le presenze degli studenti nel database SQLite.
    """
    import sqlite3
    
    try:
        # Controlla se l'emergenza è attiva
        emergenza_status = _get_emergenza_status()
        if not emergenza_status["active"]:
            return jsonify({
                "error": "Nessuna emergenza in corso",
                "message": "Non puoi compilare la rotta emergenze. Avvia prima un'emergenza."
            }), 400
        
        data = request.get_json()
        
        if not data or "classe" not in data or "presenze" not in data:
            return jsonify({"error": "Dati mancanti: classe e presenze sono obbligatori"}), 400
        
        classe = data["classe"]
        presenze = data["presenze"]
        studenti_attesi = data.get("studenti_attesi", [])
        stati_validi = {"PRESENTE", "ASSENTE", "DISPERSO"}
        
        if not isinstance(presenze, dict) or len(presenze) == 0:
            return jsonify({"error": "Nessuno studente presente nella classe"}), 400
        
        studenti_mancanti = []
        for nome_studente in studenti_attesi:
            if nome_studente not in presenze:
                studenti_mancanti.append(nome_studente)
            elif not presenze.get(nome_studente) or str(presenze.get(nome_studente)).strip() == "":
                studenti_mancanti.append(nome_studente)
        
        if studenti_mancanti:
            return jsonify({
                "error": "Compilazione incompleta",
                "message": f"I seguenti studenti non hanno uno stato assegnato: {', '.join(studenti_mancanti)}",
                "studenti": studenti_mancanti
            }), 400

        stati_non_validi = []
        for nome_studente, stato in presenze.items():
            stato_norm = str(stato).strip().upper()
            if stato_norm not in stati_validi:
                stati_non_validi.append(nome_studente)

        if stati_non_validi:
            return jsonify({
                "error": "Stati non validi",
                "message": f"Gli stati di questi studenti non sono validi: {', '.join(stati_non_validi)}",
                "studenti": stati_non_validi
            }), 400
        
        now = datetime.now()
        user = session.get("user") or {}
        record = {
            "classe": classe,
            "data": now.strftime("%Y-%m-%d"),
            "ora": now.strftime("%H:%M:%S"),
            "timestamp": now.isoformat(),
            "presenze": presenze,
            "docente_email": user.get("email", ""),
        }
        
        file_path = os.path.join(os.path.dirname(__file__), "presenze.json")
        if os.path.exists(file_path):
            with open(file_path, "r", encoding="utf-8") as f:
                try:
                    all_presenze = json.load(f)
                except json.JSONDecodeError:
                    all_presenze = []
        else:
            all_presenze = []
        
        all_presenze.append(record)
        
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(all_presenze, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            "success": True,
            "message": f"Presenze salvate per la classe {classe}",
            "record": record
        }), 200
        
    except Exception as e:
        return jsonify({"error": f"Errore durante il salvataggio: {str(e)}"}), 500


@app.route("/api/emergenze/avvia", methods=["POST"])
@role_required(['ufficio_tecnico'])  # Solo Ufficio Tecnico
def avvia_emergenza():
    """
    API per avviare un'emergenza. Solo RSPP può accedere.
    """
    try:
        status = _set_emergenza_status(active=True)
        return jsonify({
            "success": True,
            "message": "Emergenza avviata",
            "status": status
        }), 200
    except Exception as e:
        return jsonify({"error": f"Errore durante l'avvio dell'emergenza: {str(e)}"}), 500


@app.route("/api/emergenze/termina", methods=["POST"])
@role_required(['ufficio_tecnico'])  # Solo Ufficio Tecnico
def termina_emergenza():
    """
    API per terminare un'emergenza. Solo RSPP può accedere.
    """
    try:
        status = _set_emergenza_status(active=False)
        return jsonify({
            "success": True,
            "message": "Emergenza terminata",
            "status": status
        }), 200
    except Exception as e:
        return jsonify({"error": f"Errore durante la terminazione dell'emergenza: {str(e)}"}), 500

@app.route("/dashboard")
@sso_middleware.sso_login_required
def dashboard():
    """Dashboard - pagina principale per utenti autenticati."""
    if "user" not in session:
        return redirect(url_for("home"))
    user = session["user"]
    return render_template("dashboard.html", user=user)


# ============================================================
# AVVIO APPLICAZIONE
# ============================================================

with app.app_context():
    _refresh_cache_in_background()

if __name__ == "__main__":
    from config import PORT
    _refresh_cache_in_background()  # Precarica i dati appena il server parte
    app.run(host='0.0.0.0', port=PORT, debug=DEBUG)