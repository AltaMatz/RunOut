import json
from openpyxl import load_workbook # type: ignore

#SCRITTURA NEL FILE JSON DELLE AULE RAGGRUPPATE PER PUNTI DI RACCOLTA
def xlsx_to_json_by_point(puntiraccolta, output_file):
    wb = load_workbook(puntiraccolta)
    ws = wb.active

    #INTESTAZIONI
    col_headers = [cell.value for cell in ws[1]]
    row_headers = [row[0].value for row in ws.iter_rows(min_row=2, min_col=1, max_col=7)]

    points = {}
    for c_idx, col_header in enumerate(col_headers[1:], start=2):
        if col_header is None:
            continue
        point_name = ' '.join(str(col_header).split())
        points[point_name] = []
        for r_idx, row_header in enumerate(row_headers, start=2):
            cell_value = ws.cell(row=r_idx, column=c_idx).value
            if cell_value is None:
                parsed = []
            else:
                parsed = [line.strip() for line in str(cell_value).split("\n") if line.strip()]
            points[point_name].extend(parsed)

    #AGGIORNAMENTO FILE JSON
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(points, f, ensure_ascii=False, indent=4)
    print(f"File JSON aggiornato: {output_file}")

#SCRITTURA NEL FILE JSON DELLE AULE RAGGRUPPATE PER PIANO
def xlsx_to_json_by_floors(puntiraccolta, output_file):
    wb = load_workbook(puntiraccolta)
    ws = wb.active

    # Raccogli aule per piano (floor), ignorando la prima riga
    floors = {}
    
    # Leggi le righe partendo dalla seconda (ignora prima riga)
    for r_idx in range(2, ws.max_row + 1):
        floor_name = ws.cell(row=r_idx, column=1).value
        if floor_name is None:
            continue
        
        # Pulisci il nome del piano
        floor_name = ' '.join(str(floor_name).split())
        if floor_name not in floors:
            floors[floor_name] = []
        
        # Raccogli tutte le aule da questa riga (colonne 2 a 7)
        for c_idx in range(2, 8):
            cell_value = ws.cell(row=r_idx, column=c_idx).value
            if cell_value is None:
                parsed = []
            else:
                parsed = [line.strip() for line in str(cell_value).split("\n") if line.strip()]
            floors[floor_name].extend(parsed)

    #AGGIORNAMENTO FILE JSON
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(floors, f, ensure_ascii=False, indent=2)
    print(f"File JSON aggiornato: {output_file}")

if __name__ == "__main__":
    xlsx_to_json_by_point("puntiraccolta.xlsx", "puntiraccolta.json")
    xlsx_to_json_by_floors("puntiraccolta.xlsx", "floors.json")